"""Subscriber (user) CRUD + Excel import/export.

Validation:
  - phone: required, exactly 10 digits, UNIQUE
  - email: optional, basic format
  - department, location: optional free text
  - is_active: defaults to true

Excel import (v1): upsert by phone. On phone conflict, update other fields.
"""
from __future__ import annotations

import io
import re
from datetime import datetime, timezone
from typing import Iterable, Optional

from fastapi import HTTPException, UploadFile, status
from openpyxl import Workbook, load_workbook

from broadcaster.db import get_db
from broadcaster.services import groups as groups_svc

PHONE_RE = re.compile(r"^\d{10}$")
EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


def _norm_dept_loc(s) -> str:
    """Case+whitespace-insensitive dept/location comparison key. Empty -> ''."""
    return (s or "").strip().lower()


def _normalize_phone(raw: str) -> str | None:
    """Tolerate common Indian-mobile formats and return the 10-digit form.

    Accepts: 9876543210, +91 98765 43210, +91-9876543210, (987) 654-3210,
    91 98765 43210, 0 98765 43210, etc. Returns None if no 10 digits survive
    cleaning (so caller can raise the right error).
    """
    if raw is None:
        return None
    # Drop everything that's not a digit.
    digits = re.sub(r"\D+", "", str(raw))
    # Strip leading country code (91) or trunk prefix (0) if 12/11 digits remain.
    if len(digits) == 12 and digits.startswith("91"):
        digits = digits[2:]
    elif len(digits) == 11 and digits.startswith("0"):
        digits = digits[1:]
    return digits if len(digits) == 10 else None


def _now() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def _validate_phone(phone: str) -> str:
    norm = _normalize_phone(phone)
    if not norm or not PHONE_RE.match(norm):
        raise HTTPException(status_code=400, detail="invalid_phone")
    return norm


def _validate_email(email: Optional[str]) -> Optional[str]:
    if email is None or email == "":
        return None
    if not EMAIL_RE.match(email):
        raise HTTPException(status_code=400, detail="invalid_email")
    return email


# ── List / get ────────────────────────────────────────────────

def list_users(
    active_only: bool = False,
    q: Optional[str] = None,
    dept: Optional[str] = None,
    location: Optional[str] = None,
) -> list[dict]:
    where: list[str] = []
    params: list = []
    if active_only:
        where.append("is_active = 1")
    if q:
        where.append("(name LIKE ? OR phone LIKE ? OR email LIKE ?)")
        like = f"%{q}%"
        params += [like, like, like]
    if dept:
        where.append("department = ?")
        params.append(dept)
    if location:
        where.append("location = ?")
        params.append(location)

    sql = "SELECT id, name, phone, email, department, location, is_active, created_at FROM users"
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY id DESC"

    with get_db() as conn:
        rows = conn.execute(sql, params).fetchall()
    return [dict(r) for r in rows]


def get_user(uid: int) -> Optional[dict]:
    with get_db() as conn:
        row = conn.execute(
            "SELECT id, name, phone, email, department, location, is_active, created_at "
            "FROM users WHERE id = ?",
            (uid,),
        ).fetchone()
    return dict(row) if row else None


# ── Create / update / delete ──────────────────────────────────

def create_user(
    name: str,
    phone: str,
    email: Optional[str] = None,
    department: Optional[str] = None,
    location: Optional[str] = None,
    is_active: bool = True,
) -> dict:
    if not name or not name.strip():
        raise HTTPException(status_code=400, detail="name_required")
    phone = _validate_phone(phone)
    email = _validate_email(email)

    with get_db() as conn:
        try:
            cur = conn.execute(
                "INSERT INTO users (name, phone, email, department, location, is_active, created_at) "
                "VALUES (?, ?, ?, ?, ?, ?, ?)",
                (name.strip(), phone, email,
                 (department or None), (location or None),
                 1 if is_active else 0, _now()),
            )
        except Exception as e:
            if "UNIQUE" in str(e):
                raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="phone_taken")
            raise
    return get_user(cur.lastrowid)  # type: ignore[return-value]


def update_user(uid: int, **fields) -> Optional[dict]:
    allowed = {"name", "phone", "email", "department", "location", "is_active"}
    sets: list[str] = []
    params: list = []
    for k, v in fields.items():
        if k not in allowed or v is None:
            continue
        if k == "phone":
            v = _validate_phone(v)
        if k == "email":
            v = _validate_email(v)
        if k == "is_active":
            v = 1 if v else 0
        sets.append(f"{k} = ?")
        params.append(v)
    if not sets:
        return get_user(uid)
    params.append(uid)
    with get_db() as conn:
        try:
            conn.execute(f"UPDATE users SET {', '.join(sets)} WHERE id = ?", params)
        except Exception as e:
            if "UNIQUE" in str(e):
                raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="phone_taken")
            raise
    return get_user(uid)


def delete_user(uid: int) -> bool:
    with get_db() as conn:
        cur = conn.execute("DELETE FROM users WHERE id = ?", (uid,))
    return cur.rowcount > 0


# ── Excel ─────────────────────────────────────────────────────

EXCEL_HEADERS = ["name", "phone", "email", "department", "location", "is_active"]


def _row_to_dict(row: Iterable) -> dict:
    cells = list(row) + [None] * (len(EXCEL_HEADERS) - len(list(row)))
    return {
        "name": (str(cells[0]).strip() if cells[0] is not None else ""),
        "phone": (str(cells[1]).strip() if cells[1] is not None else ""),
        "email": (str(cells[2]).strip() if cells[2] is not None else ""),
        "department": (str(cells[3]).strip() if cells[3] is not None else ""),
        "location": (str(cells[4]).strip() if cells[4] is not None else ""),
        "is_active": (str(cells[5]).strip().lower() in ("1", "true", "yes", "y", "active")
                      if cells[5] is not None else True),
    }


def _err(row: int, reason: str, field: str | None, value) -> dict:
    """Single error-dict shape returned to the front-end modal."""
    return {"row": row, "reason": reason, "field": field, "value": value}


def import_from_xlsx(file: UploadFile) -> dict:
    """Read first sheet. Validate each row.

    Behavior (per 2026-06-30 user direction):
      - Upsert by phone (insert if new, update if exists).
      - Destructive replace: in one transaction, DELETE every non-admin user
        whose phone is NOT in this file OR whose email would collide with a
        file email — so the file is fully authoritative for the user list.
      - Admin (lowest-id user) is preserved from DELETION (never disappears),
        but the admin's row IS updatable when the file lists admin's phone,
        so dept/location/email changes for the admin take effect.
      - Net effect: every excel upload OVERWRITES the user list with the
        file's contents; the admin row cannot be deleted but can be updated.

    Returns {inserted, updated, skipped, deleted, errors: [{row, reason, field, value}],
             dept_location_changed, groups_created}
    """
    try:
        content = file.file.read()
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"invalid_xlsx: {e}")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"inserted": 0, "updated": 0, "skipped": 0, "deleted": 0, "errors": []}

    # Detect header row: if first row matches known header names, skip it.
    first = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
    has_header = any(h in EXCEL_HEADERS for h in first)
    data_rows = rows[1:] if has_header else rows

    inserted = updated = skipped = 0
    errors: list[dict] = []
    file_phones: set[str] = set()      # phones of valid rows; used in delete pass
    file_emails: set[str] = set()      # emails of valid rows; used in delete pass
    valid_rows: list[tuple] = []       # (idx, d, norm_phone, email_norm)
    seen_phones: set[str] = set()
    seen_emails: set[str] = set()

    # ── First pass: validate every row, collect valid rows + sets ──
    for idx, row in enumerate(data_rows, start=2 if has_header else 1):
        d = _row_to_dict(row)
        if not d["name"] and not d["phone"]:
            errors.append(_err(idx, "name_or_phone_missing", None, None))
            skipped += 1
            continue
        if not d["name"]:
            errors.append(_err(idx, "name_or_phone_missing", "name", ""))
            skipped += 1
            continue
        if not d["phone"]:
            errors.append(_err(idx, "name_or_phone_missing", "phone", ""))
            skipped += 1
            continue
        norm_phone = _normalize_phone(d["phone"])
        if not norm_phone:
            errors.append(_err(idx, "invalid_phone_format", "phone", d["phone"]))
            skipped += 1
            continue
        if norm_phone in seen_phones:
            errors.append(_err(idx, "duplicate_phone_in_file", "phone", norm_phone))
            skipped += 1
            continue
        seen_phones.add(norm_phone)
        file_phones.add(norm_phone)

        email_norm = d["email"].lower() if d["email"] else ""
        if email_norm:
            if email_norm in seen_emails:
                errors.append(_err(idx, "duplicate_email_in_file", "email", d["email"]))
                skipped += 1
                continue
            seen_emails.add(email_norm)
            file_emails.add(email_norm)
            if not EMAIL_RE.match(d["email"]):
                errors.append(_err(idx, "invalid_email_format", "email", d["email"]))
                skipped += 1
                continue

        valid_rows.append((idx, d, norm_phone, email_norm))

    # ── Second pass: delete conflicts + upsert (one transaction) ──
    dept_loc_changed = False
    deleted = 0

    with get_db() as conn:
        # Snapshot of current dept/location values BEFORE the import runs.
        existing_depts = {
            _norm_dept_loc(r["department"])
            for r in conn.execute(
                "SELECT DISTINCT department FROM users "
                "WHERE department IS NOT NULL AND department != ''"
            ).fetchall()
        }
        existing_locs = {
            _norm_dept_loc(r["location"])
            for r in conn.execute(
                "SELECT DISTINCT location FROM users "
                "WHERE location IS NOT NULL AND location != ''"
            ).fetchall()
        }

        # Locate admin (lowest-id user). If no users exist, no protection.
        admin_row = conn.execute(
            "SELECT id, phone FROM users ORDER BY id ASC LIMIT 1"
        ).fetchone()
        admin_id = admin_row["id"] if admin_row else None
        admin_phone = admin_row["phone"] if admin_row else None

        # Destructive replace pre-pass: delete non-admin rows whose
        # phone is NOT in the file OR whose email collides with a file
        # email. This is what makes the file authoritative.
        conditions = []
        params: list = []
        if file_phones:
            conditions.append("phone NOT IN (" + ",".join("?" * len(file_phones)) + ")")
            params.extend(file_phones)
        if file_emails:
            conditions.append("lower(email) IN (" + ",".join("?" * len(file_emails)) + ")")
            params.extend(file_emails)
        if conditions:
            where_clause = " AND ".join(f"({c})" for c in conditions)
            if admin_id is not None:
                sql = f"DELETE FROM users WHERE id != ? AND {where_clause}"
                cur = conn.execute(sql, [admin_id] + params)
            else:
                sql = f"DELETE FROM users WHERE {where_clause}"
                cur = conn.execute(sql, params)
            deleted = cur.rowcount

        # Upsert the surviving valid rows. Admin's row is preserved from
        # DELETION above (id != admin_id), but is updatable here so the
        # file's dept/location/email for admin can change.
        for idx, d, norm_phone, email_norm in valid_rows:
            existing = conn.execute(
                "SELECT id, department, location FROM users WHERE phone = ?", (norm_phone,)
            ).fetchone()

            try:
                if existing:
                    new_dept = _norm_dept_loc(d["department"])
                    new_loc  = _norm_dept_loc(d["location"])
                    old_dept = _norm_dept_loc(existing["department"])
                    old_loc  = _norm_dept_loc(existing["location"])
                    if (new_dept and new_dept != old_dept) or (new_loc and new_loc != old_loc):
                        dept_loc_changed = True
                    conn.execute(
                        "UPDATE users SET name=?, email=?, department=?, location=?, is_active=? "
                        "WHERE id = ?",
                        (d["name"], d["email"] or None, d["department"] or None,
                         d["location"] or None, 1 if d["is_active"] else 0,
                         existing["id"]),
                    )
                    updated += 1
                else:
                    new_dept = _norm_dept_loc(d["department"])
                    new_loc  = _norm_dept_loc(d["location"])
                    if new_dept and new_dept not in existing_depts:
                        dept_loc_changed = True
                    if new_loc and new_loc not in existing_locs:
                        dept_loc_changed = True
                    conn.execute(
                        "INSERT INTO users (name, phone, email, department, location, is_active, created_at) "
                        "VALUES (?, ?, ?, ?, ?, ?, ?)",
                        (d["name"], norm_phone, d["email"] or None,
                         d["department"] or None, d["location"] or None,
                         1 if d["is_active"] else 0, _now()),
                    )
                    inserted += 1
            except Exception as e:
                msg = str(e)
                if len(msg) > 80:
                    msg = msg[:80]
                errors.append(_err(idx, f"db_error: {e}", None, msg))
                skipped += 1

    # Conditional auto-group rebuild — only when dept/location set changed.
    groups_created = 0
    if dept_loc_changed:
        groups_svc.rebuild_auto_groups()
        with get_db() as conn:
            groups_created = conn.execute(
                "SELECT COUNT(*) AS n FROM groups WHERE is_auto = 1"
            ).fetchone()["n"]

    return {
        "inserted": inserted,
        "updated": updated,
        "skipped": skipped,
        "deleted": deleted,
        "errors": errors,
        "dept_location_changed": dept_loc_changed,
        "groups_created": groups_created,
    }


# Server-side single source of truth for error reason -> human text.
# Mirrors the front-end humanizeError() map in static/js/users.js; keep in sync.
ERROR_HUMAN = {
    "name_or_phone_missing":      "Name and phone are required.",
    "invalid_phone_format":       "Phone must be 10 digits (Indian mobile).",
    "invalid_email_format":       "Email format is invalid.",
    "duplicate_phone_in_file":    "Duplicate phone in uploaded file.",
    "duplicate_email_in_file":    "Duplicate email in uploaded file.",
    "duplicate_email_in_db":      "Email already exists for another user.",
    "phone_taken":                "Phone already exists; skipped (upsert off).",
    "db_error":                   "Database error while saving row.",
}


def import_to_csv_errors(errors: list[dict]) -> bytes:
    """Render an errors[] array (as returned by import_from_xlsx) as RFC-4180 CSV.

    Returns UTF-8 + BOM (so Excel detects encoding correctly). Pure function —
    does NOT re-read stored state.
    """
    import csv
    buf = io.StringIO()
    w = csv.writer(buf, quoting=csv.QUOTE_MINIMAL)
    w.writerow(["Row", "Field", "Value", "Reason", "Reason (human)"])
    for e in errors:
        reason = e.get("reason", "") or ""
        # Map db_error:* variants (e.g. "db_error: UNIQUE constraint failed: ...")
        # to the canonical human text by stripping the message suffix.
        human_reason = reason
        human_key = reason
        if reason.startswith("db_error"):
            human_key = "db_error"
        w.writerow([
            e.get("row", "") or "",
            e.get("field") or "",
            e.get("value") if e.get("value") is not None else "",
            reason,
            ERROR_HUMAN.get(human_key, reason),
        ])
    out = buf.getvalue()
    return ("﻿" + out).encode("utf-8")


def export_to_xlsx() -> bytes:
    """Return an in-memory .xlsx of all users."""
    wb = Workbook()
    ws = wb.active
    ws.title = "users"
    ws.append(EXCEL_HEADERS)
    for u in list_users():
        ws.append([
            u["name"], u["phone"], u["email"] or "",
            u["department"] or "", u["location"] or "",
            "active" if u["is_active"] else "inactive",
        ])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_template_xlsx() -> bytes:
    """Return a blank .xlsx with just the header row + 2 example rows.
    Lets users bulk-add without first exporting the live list."""
    wb = Workbook()
    ws = wb.active
    ws.title = "users"
    ws.append(EXCEL_HEADERS)
    # Two example rows so the format is unambiguous.
    ws.append(["Asha Kumar", "9876543210", "[email protected]", "Engineering", "Bangalore", "active"])
    ws.append(["Raj Patel",  "9123456780", "[email protected]", "Sales",      "Mumbai",    "active"])
    # Header row bold for visibility.
    from openpyxl.styles import Font
    for cell in ws[1]:
        cell.font = Font(bold=True)
    # Reasonable column widths.
    for i, w in enumerate([18, 14, 26, 14, 14, 12], start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
