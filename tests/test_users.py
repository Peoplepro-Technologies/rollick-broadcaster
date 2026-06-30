"""Phase 1b — Users CRUD + Excel."""
from __future__ import annotations

import io

import pytest
from openpyxl import Workbook


async def _login(client):
    await client.post(
        "/api/auth/login",
        data={"username": "admin", "password": "test-admin-pass"},
        headers={"Accept": "application/json"},
    )


@pytest.fixture
async def authed_client(client):
    await _login(client)
    return client


# ── Create ────────────────────────────────────────────────────

async def test_create_user_minimal(authed_client):
    r = await authed_client.post(
        "/api/users",
        json={"name": "Alice", "phone": "9876543210"},
    )
    assert r.status_code == 200
    body = r.json()
    assert body["name"] == "Alice"
    assert body["phone"] == "9876543210"
    assert body["is_active"] == 1
    assert "id" in body


async def test_create_user_full(authed_client):
    r = await authed_client.post(
        "/api/users",
        json={
            "name": "Bob",
            "phone": "1234567890",
            "email": "bob@example.com",
            "department": "Eng",
            "location": "BLR",
            "is_active": True,
        },
    )
    assert r.status_code == 200
    body = r.json()
    assert body["email"] == "bob@example.com"
    assert body["department"] == "Eng"


async def test_create_user_requires_name(authed_client):
    r = await authed_client.post("/api/users", json={"phone": "1111111111"})
    assert r.status_code == 400
    assert r.json()["detail"] == "name_and_phone_required"


async def test_create_user_rejects_short_phone(authed_client):
    r = await authed_client.post("/api/users", json={"name": "X", "phone": "12345"})
    assert r.status_code == 400
    assert r.json()["detail"] == "invalid_phone"


async def test_create_user_rejects_non_digit_phone(authed_client):
    r = await authed_client.post("/api/users", json={"name": "X", "phone": "abcdefghij"})
    assert r.status_code == 400


async def test_create_user_rejects_bad_email(authed_client):
    r = await authed_client.post(
        "/api/users", json={"name": "X", "phone": "2222222222", "email": "not-an-email"}
    )
    assert r.status_code == 400
    assert r.json()["detail"] == "invalid_email"


async def test_create_user_rejects_duplicate_phone(authed_client):
    p = {"name": "A", "phone": "3333333333"}
    r1 = await authed_client.post("/api/users", json=p)
    assert r1.status_code == 200
    r2 = await authed_client.post("/api/users", json={**p, "name": "B"})
    assert r2.status_code == 409
    assert r2.json()["detail"] == "phone_taken"


# ── Phone normalization ─────────────────────────────────────

@pytest.mark.parametrize("raw,expected", [
    ("9876543210",   "9876543210"),   # raw 10 digits
    ("+91 98765 43210", "9876543210"),  # +91 + spaces
    ("+91-9876543210",  "9876543210"),  # +91 + dashes
    ("919876543210",    "9876543210"),  # +91 no separator
    ("09876543210",     "9876543210"),  # leading 0
    ("(987) 654-3210",  "9876543210"),  # parens + dashes
    ("  +91 98765 43210  ", "9876543210"),  # whitespace
])
async def test_create_user_normalizes_phone(authed_client, raw, expected):
    r = await authed_client.post("/api/users", json={"name": f"User-{expected}", "phone": raw})
    assert r.status_code == 200, r.text
    assert r.json()["phone"] == expected


@pytest.mark.parametrize("raw", ["abc", "12345", "98765", "+91 12", "12"])
async def test_create_user_rejects_unparseable_phone(authed_client, raw):
    r = await authed_client.post("/api/users", json={"name": "X", "phone": raw})
    assert r.status_code == 400
    assert r.json()["detail"] == "invalid_phone"


async def test_excel_import_normalizes_phones(authed_client):
    """Real-world xlsx with +91 prefixes must import successfully."""
    blob = _xlsx_bytes([
        ["name", "phone", "email", "department", "location", "is_active"],
        ["Im1", "+91 98765 43210", "i1@x.com", "Eng", "BLR", "active"],
        ["Im2", "919876543211",    "i2@x.com", "Eng", "BLR", "active"],
        ["Im3", "(987) 654-3212",  "i3@x.com", "Eng", "BLR", "active"],
    ])
    files = {"file": ("users.xlsx", blob, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r = await authed_client.post("/api/users/upload-excel", files=files)
    assert r.status_code == 200
    body = r.json()
    assert body["inserted"] == 3
    assert body["skipped"] == 0
    assert body["errors"] == []

    # Verify the phones were stored normalised.
    r = await authed_client.get("/api/users", params={"q": "Im"})
    phones = sorted(u["phone"] for u in r.json())
    assert phones == ["9876543210", "9876543211", "9876543212"]

    # Cleanup.
    for u in r.json():
        await authed_client.delete(f"/api/users/{u['id']}")


# ── List / get ────────────────────────────────────────────────

async def test_list_users_returns_created(authed_client):
    for n, p in [("A", "4000000001"), ("B", "4000000002")]:
        await authed_client.post("/api/users", json={"name": n, "phone": p})
    r = await authed_client.get("/api/users")
    assert r.status_code == 200
    body = r.json()
    phones = {u["phone"] for u in body}
    assert {"4000000001", "4000000002"}.issubset(phones)


async def test_list_users_search_by_name(authed_client):
    await authed_client.post("/api/users", json={"name": "Alice", "phone": "5000000001"})
    await authed_client.post("/api/users", json={"name": "Bob", "phone": "5000000002"})
    r = await authed_client.get("/api/users", params={"q": "ali"})
    body = r.json()
    assert len(body) == 1
    assert body[0]["name"] == "Alice"


async def test_list_users_filter_active(authed_client):
    await authed_client.post("/api/users", json={"name": "On", "phone": "6000000001", "is_active": True})
    await authed_client.post("/api/users", json={"name": "Off", "phone": "6000000002", "is_active": False})
    r = await authed_client.get("/api/users", params={"active_only": "true"})
    names = {u["name"] for u in r.json()}
    assert "On" in names
    assert "Off" not in names


# ── Update / delete ───────────────────────────────────────────

async def test_update_user(authed_client):
    cr = await authed_client.post("/api/users", json={"name": "X", "phone": "7000000001"})
    uid = cr.json()["id"]
    r = await authed_client.patch(f"/api/users/{uid}", json={"department": "Sales"})
    assert r.status_code == 200
    assert r.json()["department"] == "Sales"


async def test_update_user_to_duplicate_phone_conflicts(authed_client):
    a = await authed_client.post("/api/users", json={"name": "A", "phone": "8000000001"})
    b = await authed_client.post("/api/users", json={"name": "B", "phone": "8000000002"})
    r = await authed_client.patch(f"/api/users/{b.json()['id']}", json={"phone": "8000000001"})
    assert r.status_code == 409


async def test_delete_user(authed_client):
    cr = await authed_client.post("/api/users", json={"name": "X", "phone": "9000000001"})
    uid = cr.json()["id"]
    r = await authed_client.delete(f"/api/users/{uid}")
    assert r.status_code == 200
    r2 = await authed_client.get(f"/api/users/{uid}")
    assert r2.status_code == 404


# ── Auth gating ───────────────────────────────────────────────

async def test_users_endpoints_require_auth(client):
    r = await client.get("/api/users")
    assert r.status_code == 401
    r = await client.post("/api/users", json={"name": "X", "phone": "1234567890"})
    assert r.status_code == 401


# ── Excel ─────────────────────────────────────────────────────

def _xlsx_bytes(rows: list[list]) -> bytes:
    wb = Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def test_excel_import_upserts(authed_client):
    blob = _xlsx_bytes([
        ["name", "phone", "email", "department", "location", "is_active"],
        ["Imp1", "1111111111", "i1@x.com", "Eng", "BLR", "active"],
        ["Imp2", "2222222222", "", "Sales", "MUM", "1"],
    ])
    files = {"file": ("users.xlsx", blob, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r = await authed_client.post("/api/users/upload-excel", files=files)
    assert r.status_code == 200
    body = r.json()
    assert body["inserted"] == 2
    assert body["updated"] == 0

    # Re-upload with one phone changed: should update one, insert one new.
    blob2 = _xlsx_bytes([
        ["name", "phone", "email", "department", "location", "is_active"],
        ["Imp1-renamed", "1111111111", "i1@x.com", "Eng", "BLR", "active"],
        ["Imp3", "3333333333", "", "Ops", "DEL", "1"],
    ])
    files2 = {"file": ("users.xlsx", blob2, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r2 = await authed_client.post("/api/users/upload-excel", files=files2)
    body2 = r2.json()
    assert body2["updated"] == 1
    assert body2["inserted"] == 1


async def test_excel_import_reports_invalid_rows(authed_client):
    blob = _xlsx_bytes([
        ["name", "phone"],
        ["", "12345"],                # missing name + bad phone
        ["OK", "12345"],              # bad phone
        ["OK2", "9999999999", "bad"], # too many cells is fine; this is actually valid
    ])
    # Fix last row: invalid email
    blob = _xlsx_bytes([
        ["name", "phone", "email"],
        ["", "12345"],
        ["OK", "12345"],
        ["OK2", "9999999999", "not-an-email"],
    ])
    files = {"file": ("users.xlsx", blob, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r = await authed_client.post("/api/users/upload-excel", files=files)
    body = r.json()
    assert body["inserted"] == 0
    assert body["skipped"] == 3
    reasons = {e["reason"] for e in body["errors"]}
    fields = {(e["row"], e["field"]) for e in body["errors"]}
    assert "name_or_phone_missing" in reasons
    assert "invalid_phone_format" in reasons
    assert "invalid_email_format" in reasons
    # Each error dict must carry field + value
    for e in body["errors"]:
        assert "field" in e
        assert "value" in e


async def test_excel_import_flags_in_file_dup_phone(authed_client):
    """Two rows with the same normalized phone both surface errors; no inserts."""
    blob = _xlsx_bytes([
        ["name", "phone", "email"],
        ["DupA", "9876543210", "a@x.com"],
        ["DupB", "+91 98765 43210", "b@x.com"],
    ])
    files = {"file": ("u.xlsx", blob, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r = await authed_client.post("/api/users/upload-excel", files=files)
    assert r.status_code == 200
    body = r.json()
    assert body["inserted"] == 0
    reasons = [e["reason"] for e in body["errors"]]
    assert reasons.count("duplicate_phone_in_file") == 2
    # No DB writes:
    rs = await authed_client.get("/api/users")
    assert rs.json() == []


async def test_excel_import_flags_in_file_dup_email(authed_client):
    blob = _xlsx_bytes([
        ["name", "phone", "email"],
        ["A", "1111111111", "shared@x.com"],
        ["B", "2222222222", "shared@x.com"],
    ])
    files = {"file": ("u.xlsx", blob, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r = await authed_client.post("/api/users/upload-excel", files=files)
    body = r.json()
    assert body["inserted"] == 0
    reasons = [e["reason"] for e in body["errors"]]
    assert reasons.count("duplicate_email_in_file") == 2


async def test_excel_import_flags_db_email_conflict(authed_client):
    """If email already belongs to another user, the imported row is skipped."""
    await authed_client.post(
        "/api/users", json={"name": "Owner", "phone": "5555555555", "email": "taken@x.com"}
    )
    blob = _xlsx_bytes([
        ["name", "phone", "email"],
        ["New", "6666666666", "taken@x.com"],
    ])
    files = {"file": ("u.xlsx", blob, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r = await authed_client.post("/api/users/upload-excel", files=files)
    body = r.json()
    assert body["inserted"] == 0
    assert body["errors"][0]["reason"] == "duplicate_email_in_db"


async def test_excel_import_triggers_rebuild_only_when_dept_changes(authed_client):
    """First import with a new dept should rebuild; second import with same dept should NOT."""
    blob1 = _xlsx_bytes([
        ["name", "phone", "department", "location"],
        ["U1", "1111111111", "Eng", "BLR"],
    ])
    files1 = {"file": ("u1.xlsx", blob1, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r1 = await authed_client.post("/api/users/upload-excel", files=files1)
    b1 = r1.json()
    assert b1["dept_location_changed"] is True
    assert b1["groups_created"] >= 1

    blob2 = _xlsx_bytes([
        ["name", "phone", "department", "location"],
        ["U2", "2222222222", "Eng", "BLR"],
    ])
    files2 = {"file": ("u2.xlsx", blob2, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r2 = await authed_client.post("/api/users/upload-excel", files=files2)
    b2 = r2.json()
    assert b2["dept_location_changed"] is False
    assert b2["groups_created"] == 0


async def test_excel_import_rebuild_fires_on_case_insensitive_new_dept(authed_client):
    """'Eng' and 'eng' are the same dept after lower(trim()) — second import must NOT trigger rebuild."""
    blob1 = _xlsx_bytes([
        ["name", "phone", "department"],
        ["U1", "1111111111", "Eng"],
    ])
    files1 = {"file": ("u1.xlsx", blob1, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r1 = await authed_client.post("/api/users/upload-excel", files=files1)
    assert r1.json()["dept_location_changed"] is True

    blob2 = _xlsx_bytes([
        ["name", "phone", "department"],
        ["U2", "2222222222", "eng"],  # case-only difference
    ])
    files2 = {"file": ("u2.xlsx", blob2, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r2 = await authed_client.post("/api/users/upload-excel", files=files2)
    assert r2.json()["dept_location_changed"] is False


async def test_excel_import_rebuild_detects_changed_dept(authed_client):
    """Updating an existing user's dept to a NEW value must trigger rebuild."""
    blob1 = _xlsx_bytes([
        ["name", "phone", "department"],
        ["U1", "1111111111", "Eng"],
    ])
    files1 = {"file": ("u1.xlsx", blob1, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r1 = await authed_client.post("/api/users/upload-excel", files=files1)
    assert r1.json()["dept_location_changed"] is True

    blob2 = _xlsx_bytes([
        ["name", "phone", "department"],
        ["U1", "1111111111", "Sales"],
    ])
    files2 = {"file": ("u2.xlsx", blob2, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r2 = await authed_client.post("/api/users/upload-excel", files=files2)
    b2 = r2.json()
    assert b2["updated"] == 1
    assert b2["dept_location_changed"] is True
    assert b2["groups_created"] >= 1


async def test_excel_import_rebuild_ignores_empty_dept(authed_client):
    """Empty dept strings must NOT trigger rebuild when DB had no dept."""
    blob = _xlsx_bytes([
        ["name", "phone", "department"],
        ["U1", "1111111111", ""],
    ])
    files = {"file": ("u.xlsx", blob, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r = await authed_client.post("/api/users/upload-excel", files=files)
    assert r.json()["dept_location_changed"] is False


def test_import_to_csv_errors_basic():
    from broadcaster.services.users import import_to_csv_errors
    errs = [
        {"row": 3, "field": "email",  "value": "bad@",     "reason": "invalid_email_format"},
        {"row": 5, "field": "phone",  "value": "1234",      "reason": "invalid_phone_format"},
        {"row": 9, "field": "email",  "value": "x@y.com",   "reason": "duplicate_email_in_file"},
    ]
    csv_bytes = import_to_csv_errors(errs)
    text = csv_bytes.decode("utf-8-sig")  # tolerate BOM
    lines = [ln for ln in text.splitlines() if ln]
    assert lines[0] == "Row,Field,Value,Reason,Reason (human)"
    assert "invalid_phone_format" in lines[2]
    assert "Phone must be 10 digits" in lines[2]
    assert len(lines) == 4  # header + 3 rows


def test_import_to_csv_errors_empty():
    from broadcaster.services.users import import_to_csv_errors
    csv_bytes = import_to_csv_errors([])
    text = csv_bytes.decode("utf-8-sig")
    assert text.strip().splitlines() == ["Row,Field,Value,Reason,Reason (human)"]


def test_import_to_csv_errors_handles_missing_keys():
    from broadcaster.services.users import import_to_csv_errors
    errs = [{"reason": "db_error: UNIQUE constraint failed: users.phone"}]
    out = import_to_csv_errors(errs).decode("utf-8-sig")
    assert "db_error" in out
    assert "Database error" in out


async def test_excel_errors_csv_endpoint(authed_client):
    errs = [
        {"row": 3, "field": "email", "value": "bad@", "reason": "invalid_email_format"},
        {"row": 7, "field": "phone", "value": "12345", "reason": "invalid_phone_format"},
    ]
    r = await authed_client.post("/api/users/upload-excel/errors.csv", json={"errors": errs})
    assert r.status_code == 200
    assert r.headers["content-type"].startswith("text/csv")
    assert "attachment" in r.headers["content-disposition"]
    body = r.content.decode("utf-8-sig")
    assert "invalid_email_format" in body
    assert "Email format is invalid." in body


async def test_excel_errors_csv_endpoint_rejects_empty(authed_client):
    r = await authed_client.post("/api/users/upload-excel/errors.csv", json={"errors": []})
    assert r.status_code == 400
    assert r.json()["detail"] == "no_errors"


async def test_excel_errors_csv_endpoint_unauth(client):
    r = await client.post("/api/users/upload-excel/errors.csv", json={"errors": [{"row": 1}]})
    assert r.status_code == 401


# ── Replace mode (always-on) ──────────────────────────────────

async def test_excel_import_replace_deletes_users_not_in_file(authed_client):
    """Users whose phone is NOT in the new file get deleted; admin is preserved."""
    # Seed three users (admin is auto-created with id=1; assume its phone is some
    # 10-digit number we don't know). Use the admin's row separately by id.
    await authed_client.post("/api/users", json={"name": "Keep1", "phone": "3000000001"})
    await authed_client.post("/api/users", json={"name": "Keep2", "phone": "3000000002"})
    await authed_client.post("/api/users", json={"name": "DropMe", "phone": "3000000099"})

    # Look up admin id so we can verify it's NOT in the new file but still survives.
    rs = await authed_client.get("/api/users")
    admin_id = next(u["id"] for u in rs.json() if u["name"] == "admin" or u["id"] == 1)
    # (conftest bootstrap_admin creates a user named "admin" with id=1.)

    # Upload a file that re-creates Keep1 + Keep2 but not DropMe, and also
    # adds a new user Keep3.
    blob = _xlsx_bytes([
        ["name", "phone", "department", "location"],
        ["Keep1", "3000000001", "Eng", "BLR"],
        ["Keep2", "3000000002", "Sales", "MUM"],
        ["Keep3", "3000000003", "Ops",   "DEL"],
    ])
    files = {"file": ("u.xlsx", blob, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r = await authed_client.post("/api/users/upload-excel", files=files)
    body = r.json()
    assert body["inserted"] == 1   # Keep3
    assert body["updated"] == 2    # Keep1, Keep2 (dept/loc)
    assert body["deleted"] == 1    # DropMe

    # Verify final state.
    rs = await authed_client.get("/api/users")
    names = {u["name"] for u in rs.json()}
    assert "DropMe" not in names
    assert "Keep1" in names and "Keep2" in names and "Keep3" in names
    # Admin must still be there even though the file doesn't list it.
    assert any(u["id"] == admin_id for u in rs.json())


async def test_excel_import_replace_admin_preserved_when_not_in_file(authed_client):
    """Admin's row survives a replace even when the admin phone is absent from the file."""
    rs = await authed_client.get("/api/users")
    admin_before = next(u for u in rs.json() if u["name"] == "admin" or u["id"] == 1)

    blob = _xlsx_bytes([
        ["name", "phone", "department"],
        ["OnlyOne", "4555555555", "Eng"],
    ])
    files = {"file": ("u.xlsx", blob, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    body = (await authed_client.post("/api/users/upload-excel", files=files)).json()
    assert body["inserted"] == 1

    rs = await authed_client.get("/api/users")
    names = {u["name"] for u in rs.json()}
    # Admin still there.
    assert any(u["id"] == admin_before["id"] for u in rs.json())
    assert "OnlyOne" in names


async def test_excel_import_replace_empty_file_is_noop(authed_client):
    """Header-only file doesn't delete anyone (no data rows; no deletions)."""
    await authed_client.post("/api/users", json={"name": "PreExisting", "phone": "5000000001"})
    rs = await authed_client.get("/api/users")
    n_before = len(rs.json())

    blob = _xlsx_bytes([["name", "phone"]])  # header only
    files = {"file": ("u.xlsx", blob, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    body = (await authed_client.post("/api/users/upload-excel", files=files)).json()
    assert body["inserted"] == 0 and body["updated"] == 0
    assert body["deleted"] == 0

    rs = await authed_client.get("/api/users")
    assert len(rs.json()) == n_before  # no one removed


async def test_excel_import_replace_pre_deletes_conflicting_email_user(authed_client):
    """Destructive replace: file's email takes precedence. The existing user
    who owns that email (with a DIFFERENT phone) gets deleted, then the new row
    inserts cleanly. So re-uploading the same file produces no false duplicates.
    """
    # Seed a user with email that will appear in the next upload.
    await authed_client.post(
        "/api/users", json={"name": "Old Owner", "phone": "7100000001", "email": "shared@x.com"}
    )
    blob = _xlsx_bytes([
        ["name", "phone", "email"],
        ["New Owner", "7200000002", "shared@x.com"],
    ])
    files = {"file": ("u.xlsx", blob, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    body = (await authed_client.post("/api/users/upload-excel", files=files)).json()
    assert body["inserted"] == 1
    assert body["skipped"] == 0  # no duplicate_email_in_db — pre-delete handled it
    assert body["deleted"] == 1  # Old Owner was deleted (phone not in file)

    rs = await authed_client.get("/api/users")
    phones = {u["phone"] for u in rs.json()}
    assert "7100000001" not in phones  # Old Owner gone
    assert "7200000002" in phones      # New Owner in


async def test_excel_import_replace_admin_row_in_file_is_skipped(authed_client):
    """File row matching admin's phone is reported as admin_protected and skipped."""
    # Look up the bootstrap admin row created by conftest.
    rs = await authed_client.get("/api/users")
    admin = next(u for u in rs.json() if u["name"] == "admin" or u["id"] == 1)
    admin_phone = admin["phone"]
    admin_email_before = admin.get("email") or ""
    admin_name_before = admin["name"]

    blob = _xlsx_bytes([
        ["name", "phone", "email"],
        ["Imposter", admin_phone, "evil@x.com"],
    ])
    files = {"file": ("u.xlsx", blob, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    body = (await authed_client.post("/api/users/upload-excel", files=files)).json()
    # The imposter row was skipped (admin_protected), so no inserts/updates.
    assert body["inserted"] == 0 and body["updated"] == 0
    reasons = {e["reason"] for e in body["errors"]}
    assert "admin_protected" in reasons

    # Admin row is unchanged.
    rs = await authed_client.get("/api/users")
    same_admin = next(u for u in rs.json() if u["id"] == admin["id"])
    assert same_admin["name"] == admin_name_before
    assert same_admin["email"] == admin_email_before
    assert same_admin["phone"] == admin_phone


async def test_excel_export(authed_client):
    await authed_client.post("/api/users", json={"name": "Exp", "phone": "1212121212"})
    r = await authed_client.get("/api/users/download")
    assert r.status_code == 200
    assert r.headers["content-type"].startswith("application/vnd.openxmlformats")
    assert "attachment" in r.headers["content-disposition"]
    # The body is a binary .xlsx (zip-compressed); parse it to verify content.
    import io
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(r.content), read_only=True)
    ws = wb.active
    rows = [[c.value for c in r] for r in ws.iter_rows()]
    assert any(row[0] == "Exp" and row[1] == "1212121212" for row in rows)


async def test_excel_template(authed_client):
    """Blank template has only headers + example rows, no live users."""
    # Seed a user that should NOT appear in the template.
    await authed_client.post("/api/users", json={"name": "LiveUser", "phone": "1313131313"})
    r = await authed_client.get("/api/users/template")
    assert r.status_code == 200
    assert "users_template.xlsx" in r.headers["content-disposition"]
    import io
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(r.content), read_only=True)
    ws = wb.active
    rows = [[c.value for c in r] for r in ws.iter_rows()]
    # Header row + 2 example rows only — no live users.
    assert rows[0] == ["name", "phone", "email", "department", "location", "is_active"]
    assert len(rows) == 3
    # The live user must not be present.
    assert not any(row[0] == "LiveUser" for row in rows)
    # Clean up the seed.
    r = await authed_client.get("/api/users")
    for u in r.json():
        if u["name"] == "LiveUser":
            await authed_client.delete(f"/api/users/{u['id']}")
